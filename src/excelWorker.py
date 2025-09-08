# read a CSV file
# after reading CSV add data validation (list dropdowns), auto population of columns using formulas or vlookup
# have the script lock all cells of excel except for two columns
# finally save the file as excel
# https://realpython.com/openpyxl-excel-spreadsheets-python/
# https://openpyxl.readthedocs.io/en/stable/
# https://openpyxl.pages.heptapod.net/openpyxl/index.html
# import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection
import csv 

class PlanShell:
    def __init__(self, name: str, pkg: str, pmtid: int, varn: int) -> None:
        self.name = name
        self.pkg = pkg
        self.pmtid = pmtid
        self.varn = varn
        pass

    def __str__(self) -> str:
        return f'{self.name}, {self.pkg}, {self.pmtid}, {self.varn}'

plan_list = [
    PlanShell("PPO plan in PPC no cap", '10', 2, 3),
    PlanShell("PPO plan in AltNet no cap", '11', 5, 6),
    PlanShell("PPO plan in Trad no cap", '12', 8, 9),
    PlanShell("HMO plan in PPC with cap", '13', 2, 3),
    PlanShell("HMO plan in AltNet with cap", '14', 2, 3),
]

class Region:
    def __init__(self, region_cd: str, region_nm: str) -> None:
        self.region_cd = region_cd
        self.region_nm = region_nm

    def __str__(self) -> str:
        return f'{self.region_cd}, {self.region_nm}'
    

region_list = [
    Region('A', 'Region A'),
    Region('B', 'Region B'),
    Region('C', 'Region C'),
    Region('D', 'Region D'),
    Region('E', 'Region E'),
    Region('F', 'Region F'),
    Region('G', 'Region G'),
    Region('H', 'Region H'),
    Region('I', 'Region I'),
    Region('J', 'Region J'),
]

def readCSV(filename: str, sheet) -> None:
    with open(filename, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            sheet.append(row)
    pass

def addLookupSheet(workbook) -> None:
    sheet = workbook.create_sheet('Lookup')
    
    # header row
    sheet['A1'] = 'Plan Name'
    sheet['B1'] = 'Package ID'
    sheet['C1'] = 'PMT ID'
    sheet['D1'] = 'Variation'
    # table data
    for i in range(len(plan_list)):
        sheet[f'A{i+2}'] = plan_list[i].name
        sheet[f'B{i+2}'] = plan_list[i].pkg
        sheet[f'C{i+2}'] = plan_list[i].pmtid
        sheet[f'D{i+2}'] = plan_list[i].varn

    sheet['F1'] = 'Region Code'
    sheet['G1'] = 'Region Name'
    for i in range(len(region_list)):
        sheet[f'F{i+2}'] = region_list[i].region_cd
        sheet[f'G{i+2}'] = region_list[i].region_nm
    
    pass

def addPlanListValidation(sheet) -> None:
    # dv = DataValidation(type="list", formula1='"PPO plan in PPC no cap", "PPO plan in AltNet no cap", "PPO plan in Trad no cap", "HMO plan in PPC with cap", "HMO plan in AltNet with cap"', allow_blank=True)
    
    dv = DataValidation(type="list", formula1='=$Lookup!$A$2:$A$6', allow_blank=True)
    dv.error ='Your entry is not in the list'
    dv.errorTitle = 'Invalid entry'
    dv.errorType = 'warning'
    # dv.prompt = 'Please select from the list'
    # dv.promptTitle = 'List Selection'
    dv.showErrorMessage = True
    dv.showInputMessage = True

    # attach data validation first to the sheet
    sheet.add_data_validation(dv)
    sheet.column_dimensions['D'].width = 30

    # attach data validation to entire column
    # dv.add(worksheet=sheet, col=4)
    dv.add('D1:D1048576')
    pass

def addPkgIDFormula(sheet) -> None:
    # perplexity is good
    max_row = sheet.max_row  # Get the maximum row number
    for row in range(2, max_row + 1):     # Usually start at row 2 to skip header
        # Assign VLOOKUP formula to column E for each row
        sheet[f'E{row}'].value = f'=VLOOKUP(D{row},Lookup!$A$2:$D${max_row},2,FALSE)'
    pass

def addPmtIDFormula(sheet) -> None:
    max_row = sheet.max_row  # Get the maximum row number
    for row in range(2, max_row + 1):     # Usually start at row 2 to skip header
        # Assign VLOOKUP formula to column E for each row
        sheet[f'F{row}'].value = f'=VLOOKUP(D{row},Lookup!$A$2:$D${max_row},3,FALSE)'
    pass

def addPublicFlagFormula(sheet) -> None:
    dv = DataValidation(type="list", formula1='"Y, N"', allow_blank=True)
    dv.error ='Your entry is not in the list'
    dv.errorTitle = 'Invalid entry'
    dv.errorType = 'warning'
    sheet.add_data_validation(dv)
    dv.add('L1:L1048576')
    pass

def protectSheet(sheet) -> None:
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.protection = Protection(locked=False)
    
    # sheet.security.lockStructure = True
    # sheet.protection.delete = True
    sheet.protection.sheet = True  # this will unlock the column 4 or D
    # sheet.protection.password = 'password'
    # sheet.protection.enable()
    pass


def writeExcel(workbook, filename: str) -> None:
    workbook.save(filename)  ## this will save, overwrite the workbook
    pass

def main():
    workbook = openpyxl.Workbook()
    
    sheet = workbook.active
    sheet.title = 'Data'
    readCSV('data.csv', sheet)
    addLookupSheet(workbook)

    addPlanListValidation(sheet)
    addPkgIDFormula(sheet)
    addPublicFlagFormula(sheet)
 
    protectSheet(sheet)
    writeExcel(workbook, 'output.xlsx')
    pass

if __name__ == '__main__':
    main()